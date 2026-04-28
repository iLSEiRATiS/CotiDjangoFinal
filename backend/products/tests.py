from decimal import Decimal
from io import BytesIO
from io import StringIO
from tempfile import TemporaryDirectory

import openpyxl
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from rest_framework.test import APIRequestFactory

from products.models import Category, Offer, Product, ProductImage
from products.product_importer import ProductXlsxImporter
from cotidjango.api_products import CategoriesListView, ProductListView
from users.models import CustomUser


class ProductXlsxImporterTests(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(
            username="tester",
            password="secret123",
            email="tester@example.com",
            approval_status="approved",
        )
        self.importer = ProductXlsxImporter(request_user=self.user, template_xlsx_path="")

    def _build_upload(self, headers, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            "productos.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_products_response_includes_all_products_without_omissions(self):
        category = Category.objects.create(nombre="Categoria Test")
        Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Uno",
            slug="producto-uno",
            precio="100.00",
            stock=0,
            activo=True,
        )
        Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Dos",
            slug="producto-dos",
            precio="250.50",
            stock=3,
            activo=False,
        )

        response = self.importer.export_products_response()

        workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(len(rows) - 1, Product.objects.count())
        self.assertEqual(rows[1][0], "Producto Uno")
        self.assertEqual(rows[1][1], 0)
        self.assertEqual(rows[1][17], "Si")
        self.assertIsNone(rows[1][4])
        self.assertEqual(rows[2][0], "Producto Dos")
        self.assertEqual(rows[2][1], 3)
        self.assertEqual(rows[2][17], "No")
        self.assertIsNone(rows[2][4])

    def test_export_products_response_uses_uploaded_files_when_image_urls_are_missing(self):
        with TemporaryDirectory() as tmpdir:
            with override_settings(MEDIA_ROOT=tmpdir):
                category = Category.objects.create(nombre="Imagenes")
                product = Product.objects.create(
                    user=self.user,
                    categoria=category,
                    nombre="Producto Con Archivo",
                    slug="producto-con-archivo",
                    precio="120.00",
                    stock=4,
                    activo=True,
                    imagen=SimpleUploadedFile("principal.jpg", b"main-image-bytes", content_type="image/jpeg"),
                )
                ProductImage.objects.create(
                    product=product,
                    image=SimpleUploadedFile("galeria.jpg", b"gallery-image-bytes", content_type="image/jpeg"),
                    order=1,
                    activo=True,
                )

                response = self.importer.export_products_response()

        workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        exported = next(row for row in data_rows if row[0] == "Producto Con Archivo")
        image_cell = exported[20]

        self.assertIn("/media/products/principal.jpg", image_cell)
        self.assertIn("/media/products/gallery/galeria.jpg", image_cell)

    def test_import_upload_updates_exact_product_when_idproduct_is_present(self):
        category = Category.objects.create(nombre="Velas")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Vela Ondulada Metalizada - Blister x6 Unidades",
            slug="vela-ondulada-metalizada-blister-x6-unidades",
            precio="100.00",
            stock=5,
            activo=True,
        )

        upload = self._build_upload(
            [
                "Nombre", "Stock", "SKU", "Precio", "Categorias", "Mostrar en tienda", "IDProduct",
            ],
            [
                [
                    "Vela Ondulada Metalizada Blister x6 Unidades",
                    9,
                    "",
                    150,
                    "Velas",
                    "Si",
                    product.id,
                ]
            ],
        )

        created, updated, errors = self.importer.import_upload(upload)
        product.refresh_from_db()

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertEqual(Product.objects.count(), 1)
        self.assertEqual(product.id, product.pk)
        self.assertEqual(product.precio, 150)
        self.assertEqual(product.stock, 9)
        self.assertEqual(product.nombre, "Vela Ondulada Metalizada Blister x6 Unidades")

    def test_import_upload_stops_if_required_columns_are_missing(self):
        upload = self._build_upload(
            ["Nombre", "Stock", "Categorias", "Mostrar en tienda"],
            [["Producto incompleto", 3, "Velas", "Si"]],
        )

        created, updated, errors = self.importer.import_upload(upload)

        self.assertEqual(created, 0)
        self.assertEqual(updated, 0)
        self.assertEqual(Product.objects.count(), 0)
        self.assertEqual(errors, ["Faltan columnas obligatorias: precio, sku"])

    def test_import_upload_rejects_duplicate_idproduct_inside_same_file(self):
        category = Category.objects.create(nombre="Cotillon")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Galera Test",
            slug="galera-test",
            precio="100.00",
            stock=2,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Categorias", "Mostrar en tienda", "IDProduct"],
            [
                ["Galera Test", 5, "", 110, "Cotillon", "Si", product.id],
                ["Galera Test", 7, "", 120, "Cotillon", "Si", product.id],
            ],
        )

        created, updated, errors = self.importer.import_upload(upload)
        product.refresh_from_db()

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(Product.objects.count(), 1)
        self.assertEqual(product.precio, 110)
        self.assertEqual(product.stock, 5)
        self.assertEqual(
            errors,
            [f"Fila 3: IDProduct {product.id} repetido en el archivo (ya apareció en la fila 2)."],
        )

    def test_import_upload_accepts_relative_media_urls_without_dropping_gallery(self):
        category = Category.objects.create(nombre="Imagenes")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Con Media",
            slug="producto-con-media",
            precio="100.00",
            stock=5,
            activo=True,
            image_url="/media/products/principal.jpg",
        )
        ProductImage.objects.create(
            product=product,
            image_url="/media/products/gallery/galeria.jpg",
            order=1,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Categorias", "Mostrar en tienda", "IDProduct", "URL IMAGENES"],
            [[
                "Producto Con Media",
                8,
                "",
                130,
                "Imagenes",
                "Si",
                product.id,
                "/media/products/principal.jpg | /media/products/gallery/galeria.jpg",
            ]],
        )

        created, updated, errors = self.importer.import_upload(upload)
        product.refresh_from_db()

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertEqual(product.image_url, "/media/products/principal.jpg")
        self.assertEqual(product.extra_images.count(), 1)
        self.assertEqual(product.extra_images.first().image_url, "/media/products/gallery/galeria.jpg")

    def test_import_upload_creates_product_offer_when_xlsx_marks_oferta_si(self):
        category = Category.objects.create(nombre="Ofertas")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Oferta XLSX",
            slug="producto-oferta-xlsx",
            precio="100.00",
            stock=5,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Precio oferta", "Oferta", "Categorias", "Mostrar en tienda", "IDProduct"],
            [[
                "Producto Oferta XLSX",
                8,
                "",
                100,
                80,
                "Si",
                "Ofertas",
                "Si",
                product.id,
            ]],
        )

        created, updated, errors = self.importer.import_upload(upload)
        product.refresh_from_db()
        offer = Offer.objects.get(producto=product, slug=f"xlsx-offer-product-{product.id}")

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertEqual(product.precio, 100)
        self.assertTrue(offer.activo)
        self.assertEqual(offer.porcentaje, 20)
        self.assertEqual(offer.precio_oferta, Decimal("80.00"))

    def test_import_upload_deactivates_xlsx_managed_offer_when_oferta_is_no(self):
        category = Category.objects.create(nombre="Sin Oferta")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Sin Oferta",
            slug="producto-sin-oferta",
            precio="100.00",
            stock=5,
            activo=True,
        )
        Offer.objects.create(
            nombre="Oferta XLSX - Producto Sin Oferta",
            slug=f"xlsx-offer-product-{product.id}",
            porcentaje="15.00",
            producto=product,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Precio oferta", "Oferta", "Categorias", "Mostrar en tienda", "IDProduct"],
            [[
                "Producto Sin Oferta",
                5,
                "",
                100,
                "",
                "No",
                "Sin Oferta",
                "Si",
                product.id,
            ]],
        )

        created, updated, errors = self.importer.import_upload(upload)
        offer = Offer.objects.get(producto=product, slug=f"xlsx-offer-product-{product.id}")

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertFalse(offer.activo)

    def test_import_upload_oferta_no_disables_manual_product_offer_too(self):
        category = Category.objects.create(nombre="Manual Offer")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Manual Offer",
            slug="producto-manual-offer",
            precio="100.00",
            stock=2,
            activo=True,
        )
        manual_offer = Offer.objects.create(
            nombre="Oferta Manual",
            slug="oferta-manual-producto",
            porcentaje="10.00",
            producto=product,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Precio oferta", "Oferta", "Categorias", "Mostrar en tienda", "IDProduct"],
            [[
                "Producto Manual Offer",
                2,
                "",
                100,
                "",
                "No",
                "Manual Offer",
                "Si",
                product.id,
            ]],
        )

        created, updated, errors = self.importer.import_upload(upload)
        manual_offer.refresh_from_db()

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertFalse(manual_offer.activo)

    def test_export_products_response_includes_precio_oferta_and_oferta_for_xlsx_offer(self):
        category = Category.objects.create(nombre="Categoria Oferta")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Exportado Oferta",
            slug="producto-exportado-oferta",
            precio="100.00",
            stock=7,
            activo=True,
        )
        Offer.objects.create(
            nombre="Oferta XLSX - Producto Exportado Oferta",
            slug=f"xlsx-offer-product-{product.id}",
            porcentaje="20.00",
            producto=product,
            activo=True,
        )

        response = self.importer.export_products_response()

        workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        exported = rows[1]

        self.assertEqual(exported[4], "Si")
        self.assertEqual(exported[5], "80")

    def test_export_products_response_ignores_manual_active_product_offer(self):
        category = Category.objects.create(nombre="Categoria Oferta Manual")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Exportado Oferta Manual",
            slug="producto-exportado-oferta-manual",
            precio="100.00",
            stock=7,
            activo=True,
        )
        Offer.objects.create(
            nombre="Oferta Manual Producto Exportado",
            slug="oferta-manual-exportada",
            porcentaje="15.00",
            producto=product,
            activo=True,
        )

        response = self.importer.export_products_response()

        workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        exported = next(row for row in rows[1:] if row[0] == "Producto Exportado Oferta Manual")

        self.assertIsNone(exported[4])
        self.assertIsNone(exported[5])

    def test_resolve_discount_for_product_uses_exact_offer_price_when_available(self):
        from cotidjango.api_common import resolve_discount_for_product

        category = Category.objects.create(nombre="Categoria Oferta Exacta")
        product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Producto Oferta Exacta",
            slug="producto-oferta-exacta",
            precio="881.82",
            stock=1,
            activo=True,
        )
        Offer.objects.create(
            nombre="Oferta XLSX - Producto Oferta Exacta",
            slug=f"xlsx-offer-product-{product.id}",
            porcentaje="9.28",
            precio_oferta="800.00",
            producto=product,
            activo=True,
        )

        discount = resolve_discount_for_product(product)

        self.assertIsNotNone(discount)
        self.assertEqual(discount["final_price"], Decimal("800.00"))
        self.assertEqual(discount["meta"]["percent"], 9.28)

    def test_import_upload_merges_ambiguous_rows_instead_of_creating_duplicates(self):
        cat_a = Category.objects.create(nombre="Categoria A")
        cat_b = Category.objects.create(nombre="Categoria B")
        old_product = Product.objects.create(
            user=self.user,
            categoria=cat_a,
            nombre="Silbato Plastico x24 Unidades",
            slug="silbato-plastico-a",
            precio="10.00",
            stock=2,
            activo=True,
        )
        new_product = Product.objects.create(
            user=self.user,
            categoria=cat_b,
            nombre="Silbato Plastico x24 Unidades",
            slug="silbato-plastico-b",
            precio="11.00",
            stock=3,
            activo=True,
            image_url="https://example.com/silbato.jpg",
        )

        upload = self._build_upload(
            [
                "Nombre", "Stock", "SKU", "Precio", "Mostrar en tienda",
            ],
            [
                [
                    "Silbato Plastico x24 Unidades",
                    7,
                    "",
                    99,
                    "Si",
                ]
            ],
        )

        created, updated, errors = self.importer.import_upload(upload)

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertEqual(Product.objects.count(), 1)

        survivor = Product.objects.get()
        self.assertEqual(survivor.pk, new_product.pk)
        self.assertNotEqual(survivor.pk, old_product.pk)
        self.assertEqual(survivor.precio, 99)
        self.assertEqual(survivor.stock, 7)
        self.assertEqual(survivor.image_url, "https://example.com/silbato.jpg")

    def test_import_upload_merges_duplicates_in_same_category_and_keeps_useful_data(self):
        category = Category.objects.create(nombre="Cotillon")
        old_product = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Galera Argentina",
            slug="galera-argentina-vieja",
            precio="100.00",
            stock=0,
            activo=False,
            descripcion="Descripcion completa",
        )
        ProductImage.objects.create(
            product=old_product,
            image_url="https://example.com/galera-extra.jpg",
            order=1,
            activo=True,
        )
        survivor = Product.objects.create(
            user=self.user,
            categoria=category,
            nombre="Galera Argentina",
            slug="galera-argentina-nueva",
            precio="120.00",
            stock=4,
            activo=True,
        )

        upload = self._build_upload(
            ["Nombre", "Stock", "SKU", "Precio", "Categorias", "Mostrar en tienda", "IDProduct"],
            [["Galera Argentina", 9, "", 150, "Cotillon", "Si", survivor.id]],
        )

        created, updated, errors = self.importer.import_upload(upload)

        self.assertEqual(created, 0)
        self.assertEqual(updated, 1)
        self.assertEqual(errors, [])
        self.assertEqual(Product.objects.count(), 1)

        survivor.refresh_from_db()
        self.assertEqual(survivor.descripcion, "Descripcion completa")
        self.assertEqual(survivor.precio, 150)
        self.assertEqual(survivor.stock, 9)
        self.assertTrue(survivor.activo)
        self.assertEqual(survivor.extra_images.count(), 1)
        self.assertEqual(survivor.extra_images.first().image_url, "https://example.com/galera-extra.jpg")


class CategoryValidationTests(TestCase):
    def test_category_cannot_be_its_own_parent(self):
        category = Category.objects.create(nombre="Velas")
        category.parent = category

        with self.assertRaises(ValidationError):
            category.full_clean()

    def test_category_cannot_use_descendant_as_parent(self):
        root = Category.objects.create(nombre="Cotillon")
        child = Category.objects.create(nombre="Vinchas", parent=root)
        grandchild = Category.objects.create(nombre="Brillos", parent=child)
        root.parent = grandchild

        with self.assertRaises(ValidationError):
            root.full_clean()


class SanitizeCategoryMovesCommandTests(TestCase):
    def setUp(self):
        self.root_bengalas = Category.objects.create(nombre="Bengalas")
        self.velas = Category.objects.create(nombre="Velas")
        self.child_bengalas = Category.objects.create(nombre="Bengalas", parent=self.velas)
        self.manualidades_target = Category.objects.create(nombre="Artículos Para Manualidades")
        self.manualidades_source = Category.objects.create(nombre="Librería y Manualidades")
        self.user = CustomUser.objects.create_user(
            username="commandtester",
            password="secret123",
            email="commandtester@example.com",
            approval_status="approved",
        )
        Product.objects.create(
            user=self.user,
            categoria=self.child_bengalas,
            nombre="Vela Bengala Test",
            slug="vela-bengala-test",
            precio="100.00",
            stock=1,
            activo=True,
        )
        Product.objects.create(
            user=self.user,
            categoria=self.manualidades_source,
            nombre="Manualidad Test",
            slug="manualidad-test",
            precio="50.00",
            stock=2,
            activo=True,
        )

    def test_sanitize_category_moves_simulation_does_not_persist_changes(self):
        out = StringIO()

        call_command("sanitize_category_moves", stdout=out)

        self.assertEqual(Product.objects.filter(categoria=self.child_bengalas).count(), 1)
        self.assertEqual(Product.objects.filter(categoria=self.root_bengalas).count(), 0)
        self.assertTrue(Category.objects.filter(pk=self.child_bengalas.pk).exists())
        self.assertIn("MODO SIMULACION", out.getvalue())
        self.assertIn("Detalle final por operacion:", out.getvalue())

    def test_sanitize_category_moves_apply_persists_expected_changes(self):
        out = StringIO()

        call_command("sanitize_category_moves", "--apply", stdout=out)

        self.assertEqual(Product.objects.filter(categoria=self.child_bengalas).count(), 0)
        self.assertEqual(Product.objects.filter(categoria=self.root_bengalas).count(), 1)
        self.assertFalse(Category.objects.filter(pk=self.child_bengalas.pk).exists())
        self.assertEqual(Product.objects.filter(categoria=self.manualidades_source).count(), 0)
        self.assertEqual(Product.objects.filter(categoria=self.manualidades_target).count(), 1)
        self.assertFalse(Category.objects.filter(pk=self.manualidades_source.pk).exists())
        self.assertIn("estado=applied", out.getvalue())


class DedupeCategoriesCommandTests(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(
            username="dedupetester",
            password="secret123",
            email="dedupetester@example.com",
            approval_status="approved",
        )
        self.velas = Category.objects.create(nombre="Velas")
        self.canonical = Category.objects.create(nombre="Velas con Luz", parent=self.velas)
        self.duplicate = Category.objects.create(nombre="Velas con Luz", parent=self.velas)
        self.child = Category.objects.create(nombre="Subcategoria", parent=self.duplicate)
        Product.objects.create(
            user=self.user,
            categoria=self.duplicate,
            nombre="Producto Duplicado",
            slug="producto-duplicado-velas",
            precio="100.00",
            stock=1,
            activo=True,
        )
        self.offer = Offer.objects.create(
            nombre="Oferta Duplicada",
            porcentaje="10.00",
            categoria=self.duplicate,
            activo=True,
        )

    def test_dedupe_categories_simulation_does_not_persist_changes(self):
        out = StringIO()

        call_command("dedupe_categories", stdout=out)

        self.assertTrue(Category.objects.filter(pk=self.duplicate.pk).exists())
        self.assertEqual(Product.objects.filter(categoria=self.duplicate).count(), 1)
        self.assertEqual(Offer.objects.filter(categoria=self.duplicate).count(), 1)
        self.assertEqual(Category.objects.filter(parent=self.duplicate).count(), 1)
        self.assertIn("MODO SIMULACION", out.getvalue())

    def test_dedupe_categories_apply_merges_duplicates(self):
        out = StringIO()

        call_command("dedupe_categories", "--apply", stdout=out)

        self.assertFalse(Category.objects.filter(pk=self.duplicate.pk).exists())
        self.assertEqual(Product.objects.filter(categoria=self.canonical).count(), 1)
        self.assertEqual(Offer.objects.filter(categoria=self.canonical).count(), 1)
        self.child.refresh_from_db()
        self.assertEqual(self.child.parent_id, self.canonical.pk)
        self.assertIn("estado=applied", out.getvalue())


class AuditCatalogXlsxCommandTests(TestCase):
    def _build_catalog_workbook(self, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([
            "Nombre", "Stock", "SKU", "Precio", "Precio oferta",
            "Nombre atributo 1", "Valor atributo 1", "Nombre atributo 2", "Valor atributo 2",
            "Nombre atributo 3", "Valor atributo 3", "Categorías", "Peso", "Alto",
            "Ancho", "Profundidad", "Mostrar en tienda", "IDProduct", "IDStock", "URL IMAGENES",
        ])
        for row in rows:
            sheet.append(row)
        return workbook

    def test_audit_catalog_xlsx_reports_clean_workbook(self):
        user = CustomUser.objects.create_user(
            username="auditclean",
            password="secret123",
            email="auditclean@example.com",
            approval_status="approved",
        )
        category = Category.objects.create(nombre="Cotillon")
        product = Product.objects.create(
            user=user,
            categoria=category,
            nombre="Producto Uno",
            slug="producto-uno-audit",
            precio="100.00",
            stock=2,
            activo=True,
        )
        workbook = self._build_catalog_workbook([
            ["Producto Uno", 2, "", 100, "", "", "", "", "", "", "", "Cotillon", "", "", "", "", "Si", product.id, "", ""],
            ["Producto Dos", 0, "", 200, "", "", "", "", "", "", "", "Velas > Velas con Luz", "", "", "", "", "Si", 99999, "", ""],
        ])

        with TemporaryDirectory() as tmpdir:
            path = f"{tmpdir}/catalogo.xlsx"
            workbook.save(path)
            out = StringIO()

            call_command("audit_catalog_xlsx", path, stdout=out)

        output = out.getvalue()
        self.assertIn("filas utiles: 2", output)
        self.assertIn("grupos duplicados en XLSX por Nombre + Categorias: 0", output)
        self.assertIn("IDs duplicados en XLSX: 0", output)
        self.assertIn("IDs presentes en el XLSX pero ausentes en DB: 1", output)

    def test_audit_catalog_xlsx_reports_duplicate_workbook_rows(self):
        workbook = self._build_catalog_workbook([
            ["Producto Uno", 2, "", 100, "", "", "", "", "", "", "", "Cotillon", "", "", "", "", "Si", 10, "", ""],
            ["Producto Uno", 2, "", 100, "", "", "", "", "", "", "", "Cotillon", "", "", "", "", "Si", 10, "", ""],
        ])

        with TemporaryDirectory() as tmpdir:
            path = f"{tmpdir}/catalogo_duplicado.xlsx"
            workbook.save(path)
            out = StringIO()

            call_command("audit_catalog_xlsx", path, stdout=out)

        output = out.getvalue()
        self.assertIn("grupos duplicados en XLSX por Nombre + Categorias: 1", output)
        self.assertIn("IDs duplicados en XLSX: 1", output)


class OffersVirtualCategoryApiTests(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.user = CustomUser.objects.create_user(
            username="offerapi",
            password="secret123",
            email="offerapi@example.com",
            approval_status="approved",
        )
        self.root = Category.objects.create(nombre="Guirnaldas y Decoración", slug="guirnaldas-y-decoracion")
        self.product = Product.objects.create(
            user=self.user,
            categoria=self.root,
            nombre="Producto con Oferta",
            slug="producto-con-oferta",
            precio="100.00",
            stock=3,
            activo=True,
        )
        Offer.objects.create(
            nombre="Oferta XLSX - Producto con Oferta",
            slug=f"xlsx-offer-product-{self.product.id}",
            porcentaje="20.00",
            precio_oferta="80.00",
            producto=self.product,
            activo=True,
        )

    def test_categories_list_includes_ofertas_root(self):
        request = self.factory.get("/api/categories-list")
        response = CategoriesListView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        names = [item["nombre"] for item in response.data["items"]]
        self.assertIn("Ofertas", names)

    def test_products_list_returns_discounted_products_for_ofertas_category(self):
        request = self.factory.get("/api/products", {"category": "ofertas", "page": 1, "limit": 24})
        response = ProductListView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["total"], 1)
        self.assertEqual(response.data["items"][0]["name"], "Producto con Oferta")
        self.assertEqual(response.data["items"][0]["priceOriginal"], 100.0)
        self.assertEqual(response.data["items"][0]["price"], 80.0)
