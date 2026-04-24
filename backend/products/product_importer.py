from decimal import Decimal
import hashlib
import os
import unicodedata

import openpyxl
from django.http import HttpResponse
from django.utils.text import slugify

from .models import Category, Offer, Product, ProductImage


PRODUCT_HEADERS = [
    "sku", "parent_sku", "nombre", "slug", "descripcion", "categoria", "subcategoria", "marca",
    "precio", "costo", "moneda", "stock", "activo", "opcion_1_nombre", "opcion_1_valor",
    "opcion_2_nombre", "opcion_2_valor", "imagen_1", "url imag", "imagen_2", "imagen_3", "imagen_4",
    "imagen_5", "meta_title",
    "meta_description", "peso", "largo", "ancho", "alto", "es_destacado", "requiere_envio",
    "gestion_stock",
]

EXPORT_HEADERS = [
    "Nombre", "Stock", "SKU", "Precio", "Precio oferta", "Nombre atributo 1", "Valor atributo 1",
    "Nombre atributo 2", "Valor atributo 2", "Nombre atributo 3", "Valor atributo 3", "Categorías",
    "Peso", "Alto", "Ancho", "Profundidad", "Mostrar en tienda", "IDProduct", "IDStock", "URL IMAGENES",
]

EXPORT_COLUMN_WIDTHS = {
    "A": 13.0,
    "B": 54.75,
    "C": 13.0,
    "D": 13.0,
    "E": 13.0,
    "F": 13.0,
    "G": 13.0,
    "H": 13.0,
    "I": 13.0,
    "J": 13.0,
    "K": 13.0,
    "L": 13.0,
    "M": 13.0,
    "N": 13.0,
    "O": 13.0,
    "P": 13.0,
    "Q": 13.0,
    "R": 13.0,
    "S": 13.0,
    "T": 13.0,
}


SAMPLE_ROWS = [
    {
        "sku": "REM-BAS-001",
        "parent_sku": "",
        "nombre": "Remera basica",
        "slug": "remera-basica",
        "descripcion": "Remera algodon lisa",
        "categoria": "Ropa",
        "subcategoria": "Remeras",
        "marca": "Acme",
        "precio": 8999,
        "costo": 4500,
        "moneda": "ARS",
        "stock": 120,
        "activo": True,
        "opcion_1_nombre": "",
        "opcion_1_valor": "",
        "opcion_2_nombre": "",
        "opcion_2_valor": "",
        "imagen_1": "https://example.com/rem-bas-001.jpg",
        "url imag": "",
        "imagen_2": "",
        "meta_title": "Remera basica",
        "meta_description": "Remera de algodon basica",
        "peso": 0.3,
        "largo": 30,
        "ancho": 25,
        "alto": 2,
        "es_destacado": True,
        "requiere_envio": True,
        "gestion_stock": True,
    },
    {
        "sku": "REM-BAS-001-M-NEGRO",
        "parent_sku": "REM-BAS-001",
        "nombre": "Remera basica M Negro",
        "slug": "remera-basica-m-negro",
        "descripcion": "Remera algodon talla M color negro",
        "categoria": "Ropa",
        "subcategoria": "Remeras",
        "marca": "Acme",
        "precio": 8999,
        "costo": 4500,
        "moneda": "ARS",
        "stock": 40,
        "activo": True,
        "opcion_1_nombre": "Talle",
        "opcion_1_valor": "M",
        "opcion_2_nombre": "Color",
        "opcion_2_valor": "Negro",
        "imagen_1": "https://example.com/rem-bas-001-m-negro.jpg",
        "url imag": "",
        "imagen_2": "",
        "meta_title": "Remera basica M negro",
        "meta_description": "Remera negra basica talla M",
        "peso": 0.3,
        "largo": 30,
        "ancho": 25,
        "alto": 2,
        "es_destacado": False,
        "requiere_envio": True,
        "gestion_stock": True,
    },
]


HEADER_ALIAS = {
    "categorias": "categoria",
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "sub categoria": "subcategoria",
    "subcategor?a": "subcategoria",
    "url imagenes": "url imag",
    "url imag": "url imag",
    "url_imag": "url imag",
    "imagen 1": "imagen_1",
    "imagen 2": "imagen_2",
    "imagen 3": "imagen_3",
    "imagen 4": "imagen_4",
    "imagen 5": "imagen_5",
    "mostrar en tienda": "activo",
    "idproduct": "idproduct",
    "id product": "idproduct",
    "id_product": "idproduct",
    "idstock": "idstock",
    "id stock": "idstock",
    "id_stock": "idstock",
    "nombre atributo 1": "opcion_1_nombre",
    "valor atributo 1": "opcion_1_valor",
    "nombre atributo1": "opcion_1_nombre",
    "valor atributo1": "opcion_1_valor",
    "nombre atributo 2": "opcion_2_nombre",
    "valor atributo 2": "opcion_2_valor",
    "nombre atributo2": "opcion_2_nombre",
    "valor atributo2": "opcion_2_valor",
    "nombre atributo 3": "opcion_3_nombre",
    "valor atributo 3": "opcion_3_valor",
    "nombre atributo3": "opcion_3_nombre",
    "valor atributo3": "opcion_3_valor",
}


class ProductXlsxImporter:
    def __init__(self, *, request_user, template_xlsx_path):
        self.request_user = request_user
        self.template_xlsx_path = template_xlsx_path

    def export_workbook(self, rows, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(PRODUCT_HEADERS)
        for row in rows:
            ws.append([row.get(header, "") for header in PRODUCT_HEADERS])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_template_response(self):
        if os.path.isfile(self.template_xlsx_path):
            with open(self.template_xlsx_path, "rb") as fh:
                response = HttpResponse(
                    fh.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = 'attachment; filename="plantilla_productos.xlsx"'
                return response
        empty_row = {header: "" for header in PRODUCT_HEADERS}
        return self.export_workbook([empty_row], "plantilla_productos.xlsx")

    def export_products_response(self):
        products = (
            Product.objects.all()
            .select_related("categoria", "categoria__parent")
            .prefetch_related("extra_images")
            .order_by("id")
        )
        # Exportar siempre desde cero evita omisiones provocadas por plantillas base
        # o deduplicaciones heurísticas. La planilla debe reflejar exactamente la DB.
        workbook = self._build_export_workbook()
        ws = workbook.active

        for product in products:
            row = self._serialize_product_for_export(product)
            ws.append([row.get(header, "") for header in EXPORT_HEADERS])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="productos_existentes.xlsx"'
        workbook.save(response)
        return response

    def import_upload(self, upload):
        created = 0
        updated = 0
        errors = []

        def open_sheet():
            upload.seek(0)
            wb_local = openpyxl.load_workbook(upload, data_only=True, read_only=True)
            return wb_local, wb_local.active

        wb, sheet = open_sheet()
        header_idx = None
        header_row = None
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if not row:
                continue
            normalized = [self._norm_header(cell) for cell in row]
            if {"sku", "nombre", "precio"}.issubset(set(normalized)):
                header_idx = i
                header_row = row
                break
        wb.close()

        if header_row is None:
            header_row = PRODUCT_HEADERS
            header_idx = -1

        headers = [self._norm_header(cell) for cell in header_row]
        header_map = {}
        for idx, header in enumerate(headers):
            key = HEADER_ALIAS.get(header, header)
            if key:
                header_map[key] = idx

        missing = {"sku", "nombre", "precio"} - set(header_map.keys())
        if missing:
            errors.append(f"Faltan columnas obligatorias: {', '.join(sorted(missing))}")

        sku_name_sets = {}
        wb, sheet = open_sheet()
        for row_idx, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_idx is not None and row_idx - 1 == header_idx:
                continue
            row_data = {
                key: raw[header_map[key]] if key in header_map and header_map[key] < len(raw) else ""
                for key in header_map
            }
            nombre_row = str(row_data.get("nombre") or "").strip()
            sku_row = str(row_data.get("sku") or "").strip().upper()
            if not nombre_row or not sku_row:
                continue
            sku_name_sets.setdefault(sku_row, set()).add(self._norm_header(nombre_row))
        wb.close()
        multi_name_skus = {sku for sku, names in sku_name_sets.items() if len(names) > 1}

        wb, sheet = open_sheet()
        for idx, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_idx is not None and idx - 1 == header_idx:
                continue
            data = {
                key: raw[header_map[key]] if key in header_map and header_map[key] < len(raw) else ""
                for key in header_map
            }
            row_data = {key: data.get(key) for key in header_map}
            if all(value in ("", None) for value in row_data.values()):
                continue

            nombre = row_data.get("nombre") or ""
            precio = self._parse_decimal(row_data.get("precio"))
            if not nombre:
                errors.append(f"Fila {idx}: nombre es obligatorio.")
                continue
            if precio is None:
                precio = Decimal("0")

            attr_pairs = self._collect_attr_pairs(row_data)
            has_declared_attrs = bool(attr_pairs)

            sku_raw = row_data.get("sku") or ""
            parent_sku_raw = row_data.get("parent_sku") or ""
            slug_raw = row_data.get("slug") or ""
            idproduct_raw = row_data.get("idproduct") or ""
            categoria_raw = row_data.get("categoria") or ""
            subcategoria_raw = row_data.get("subcategoria") or ""

            categoria_obj = None
            path_parts = self._compose_category_path(categoria_raw, subcategoria_raw)
            if path_parts:
                parent = None
                for part in path_parts:
                    parent = self._get_or_create_category_normalized(part, parent=parent)
                categoria_obj = parent

            existing, identity_error = self._find_existing_product_identity(
                idproduct_raw=idproduct_raw,
                slug_raw=slug_raw,
                nombre=nombre,
                categoria_obj=categoria_obj,
                grouped=has_declared_attrs,
            )
            if identity_error:
                errors.append(f"Fila {idx}: {identity_error}")
                continue

            if has_declared_attrs:
                slug = existing.slug if existing else self._build_group_slug(nombre=nombre, path_parts=path_parts)
            else:
                sku_upper = str(sku_raw).strip().upper()
                effective_sku = "" if (sku_upper and sku_upper in multi_name_skus) else sku_raw
                slug = (
                    existing.slug
                    if existing
                    else self._build_identity_slug(
                        sku_raw=effective_sku,
                        slug_raw=slug_raw,
                        nombre=nombre,
                        path_parts=path_parts,
                        parent_sku_raw=parent_sku_raw or sku_raw,
                    )
                )

            existing = existing or Product.objects.filter(slug=slug).first()
            is_new = existing is None
            product = existing or Product(slug=slug, user=self.request_user)
            product.nombre = nombre
            product.descripcion = row_data.get("descripcion") or ""
            product.precio = precio
            stock = self._parse_int(row_data.get("stock"))
            product.stock = stock if stock is not None else 0
            product.activo = self._parse_bool(row_data.get("activo"), default=True)
            product.categoria = categoria_obj

            attrs = {name: values for name, values in attr_pairs}

            has_sku = bool(str(sku_raw).strip())
            if has_declared_attrs:
                merged = {}
                if isinstance(product.atributos, dict):
                    merged.update(product.atributos)
                for key, values in attrs.items():
                    current = merged.get(key, [])
                    if not isinstance(current, list):
                        current = [current] if current else []
                    for value in values:
                        if value not in current:
                            current.append(value)
                    merged[key] = current
                product.atributos = merged if merged else {}

                stock_map = product.atributos_stock if isinstance(product.atributos_stock, dict) else {}
                price_map = product.atributos_precio if isinstance(product.atributos_precio, dict) else {}
                if stock is not None:
                    for key, values in attrs.items():
                        existing_map = stock_map.get(key, {}) if isinstance(stock_map.get(key), dict) else {}
                        for value in values:
                            existing_map[value] = max(int(existing_map.get(value, 0)), int(stock))
                        stock_map[key] = existing_map
                product.atributos_stock = stock_map if stock_map else {}
                for key, values in attrs.items():
                    existing_prices = price_map.get(key, {}) if isinstance(price_map.get(key), dict) else {}
                    for value in values:
                        existing_prices[value] = float(precio)
                    price_map[key] = existing_prices
                product.atributos_precio = price_map if price_map else {}
                product.precio = self._resolve_base_price(product, fallback=precio)
            elif has_sku:
                product.atributos = attrs if attrs else {}
                if stock is not None and attrs:
                    stock_map = {}
                    for key, values in attrs.items():
                        stock_map[key] = {value: int(stock) for value in values}
                    product.atributos_stock = stock_map
                elif not attrs:
                    product.atributos_stock = {}
                product.atributos_precio = {}
            else:
                product.atributos = {}
                product.atributos_stock = {}
                product.atributos_precio = {}

            image_urls = self._extract_image_urls(row_data)
            if image_urls:
                product.image_url = image_urls[0]
            if is_new and not product.slug:
                product.slug = self._build_slug(nombre)
            product.save()

            if image_urls:
                gallery_urls = image_urls[1:]
                existing_qs = ProductImage.objects.filter(product=product)
                existing_by_url = {item.image_url: item for item in existing_qs if item.image_url}
                keep = set()
                for idx2, url in enumerate(gallery_urls, start=1):
                    keep.add(url)
                    existing_image = existing_by_url.get(url)
                    if existing_image:
                        changed = False
                        if existing_image.order != idx2:
                            existing_image.order = idx2
                            changed = True
                        if not existing_image.activo:
                            existing_image.activo = True
                            changed = True
                        if changed:
                            existing_image.save(update_fields=["order", "activo"])
                    else:
                        ProductImage.objects.create(
                            product=product,
                            image_url=url,
                            order=idx2,
                            activo=True,
                        )
                existing_qs.exclude(image_url__in=keep).delete()
            else:
                ProductImage.objects.filter(product=product).delete()

            if is_new:
                created += 1
            else:
                updated += 1
            product = self._merge_same_name_duplicates(
                product=product,
                category_scoped=bool(categoria_obj),
            )

        wb.close()
        return created, updated, errors

    def _parse_bool(self, value, default=True):
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in {"true", "1", "si", "sÃƒÂ­", "yes", "y"}

    def _parse_decimal(self, value):
        if value is None or value == "":
            return None
        text = str(value).replace("$", "").replace(" ", "").replace("\xa0", "")
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return Decimal(text)
        except Exception:
            return None

    def _parse_int(self, value):
        if value in (None, ""):
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    def _parse_attr_values(self, value):
        if value in (None, ""):
            return []
        text = str(value).strip()
        if not text:
            return []
        for sep in [",", ";", "|", "/"]:
            if sep in text:
                parts = [part.strip() for part in text.split(sep)]
                return [part for part in parts if part]
        return [text]

    def _collect_attr_pairs(self, row_data):
        pairs = []
        for idx in (1, 2, 3):
            name = (row_data.get(f"opcion_{idx}_nombre") or "").strip()
            values = self._parse_attr_values(row_data.get(f"opcion_{idx}_valor"))
            if name and values:
                pairs.append((name, values))
        return pairs

    def _extract_image_urls(self, row_data):
        raw_values = [
            row_data.get("imagen_1"),
            row_data.get("url imag"),
            row_data.get("url_imag"),
            row_data.get("imagen_2"),
            row_data.get("imagen_3"),
            row_data.get("imagen_4"),
            row_data.get("imagen_5"),
        ]
        output = []
        seen = set()
        for raw in raw_values:
            if raw in (None, ""):
                continue
            text = str(raw).strip()
            if not text:
                continue
            parts = [text]
            for sep in ["|", ";", ","]:
                if sep in text:
                    parts = [part.strip() for part in text.split(sep)]
                    break
            for part in parts:
                if not part or not part.startswith(("http://", "https://")):
                    continue
                if part in seen:
                    continue
                seen.add(part)
                output.append(part)
        return output

    def _build_slug(self, base):
        candidate = slugify(base or "")[:110] or "producto"
        original = candidate
        counter = 1
        while Product.objects.filter(slug=candidate).exists():
            counter += 1
            candidate = f"{original}-{counter}"
        return candidate

    def _norm_header(self, value):
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(char for char in text if not unicodedata.combining(char))
        return text

    def _norm_compare_text(self, value):
        text = self._norm_header(value)
        text = "".join(char if char.isalnum() else " " for char in text)
        return " ".join(text.split())

    def _parse_category_path(self, raw):
        if not raw:
            return []
        text = str(raw).strip()
        if not text:
            return []
        for sep in (" > ", ">", " / ", "/", " | ", "|", " - ", " Ã¢â‚¬â€œ "):
            if sep in text:
                parts = [part.strip() for part in text.split(sep)]
                return [part for part in parts if part]
        return [text]

    def _compose_category_path(self, categoria_raw, subcategoria_raw):
        categoria_parts = self._parse_category_path(categoria_raw)
        subcategoria_parts = self._parse_category_path(subcategoria_raw)
        if not categoria_parts and not subcategoria_parts:
            return []
        if not categoria_parts:
            return subcategoria_parts
        if not subcategoria_parts:
            return categoria_parts

        categoria_norm = [self._norm_header(part) for part in categoria_parts]
        output = list(categoria_parts)
        for part in subcategoria_parts:
            if self._norm_header(part) not in categoria_norm:
                output.append(part)
        return output

    def _get_or_create_category_normalized(self, name, parent=None):
        target = self._norm_header(name)
        for category in Category.objects.filter(parent=parent):
            if self._norm_header(category.nombre) == target:
                return category
        return Category.objects.create(nombre=name, parent=parent)

    def _build_identity_slug(self, *, sku_raw, slug_raw, nombre, path_parts, parent_sku_raw):
        sku_text = str(sku_raw or "").strip()
        if sku_text:
            return slugify(sku_text)[:110] or self._build_slug(nombre)

        base_slug = slugify(slug_raw or nombre or "")[:70] or "producto"
        path_key = "|".join(self._norm_header(part) for part in (path_parts or []))
        raw = "|".join([
            self._norm_header(nombre),
            path_key,
            self._norm_header(parent_sku_raw),
        ])
        digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
        return f"{base_slug}-{digest}"[:110]

    def _find_existing_product_by_name(self, *, nombre, categoria_obj):
        target = self._norm_compare_text(nombre)
        queryset = Product.objects.all()
        if categoria_obj is None:
            queryset = queryset.filter(categoria__isnull=True)
        else:
            queryset = queryset.filter(categoria=categoria_obj)
        for product in queryset.order_by("id"):
            if self._norm_compare_text(product.nombre) == target:
                return product
        return None

    def _find_products_by_normalized_name(self, *, nombre):
        target = self._norm_compare_text(nombre)
        if not target:
            return []
        matches = []
        for product in Product.objects.all().order_by("id"):
            if self._norm_compare_text(product.nombre) == target:
                matches.append(product)
        return matches

    def _find_existing_group_product(self, *, nombre, categoria_obj):
        return self._find_existing_product_by_name(nombre=nombre, categoria_obj=categoria_obj)

    def _find_existing_product_identity(self, *, idproduct_raw, slug_raw, nombre, categoria_obj, grouped):
        pk = self._parse_int(idproduct_raw)
        if pk:
            product = Product.objects.filter(pk=pk).first()
            if product:
                return product, None
            return None, f"IDProduct {pk} no existe. Para evitar duplicados, esa fila no se importó."

        slug_text = str(slug_raw or "").strip()
        if slug_text:
            product = Product.objects.filter(slug=slug_text).first()
            if product:
                return product, None

        name_matches = self._find_products_by_normalized_name(nombre=nombre)
        if categoria_obj is not None:
            category_matches = [product for product in name_matches if product.categoria_id == categoria_obj.id]
            if len(category_matches) == 1:
                return category_matches[0], None
            if len(category_matches) > 1:
                return self._select_best_duplicate_candidate(
                    category_matches,
                    preferred_category_id=categoria_obj.id,
                ), None

        if len(name_matches) == 1:
            return name_matches[0], None
        if len(name_matches) > 1:
            return self._select_best_duplicate_candidate(
                name_matches,
                preferred_category_id=categoria_obj.id if categoria_obj else None,
            ), None

        if grouped:
            return self._find_existing_group_product(nombre=nombre, categoria_obj=categoria_obj), None
        return self._find_existing_product_by_name(nombre=nombre, categoria_obj=categoria_obj), None

    def _select_best_duplicate_candidate(self, candidates, preferred_category_id=None):
        def score(product):
            has_primary_image = bool(str(product.image_url or "").strip()) or bool(getattr(product, "imagen", None))
            has_gallery = product.extra_images.exists()
            has_description = bool(str(product.descripcion or "").strip())
            has_video = bool(str(product.video_url or "").strip())
            has_stock = int(product.stock or 0) > 0
            has_price = self._parse_decimal(product.precio) not in (None, Decimal("0"))
            has_orders = product.order_items.exists()
            has_offers = product.ofertas.exists()
            return (
                1 if preferred_category_id and product.categoria_id == preferred_category_id else 0,
                1 if has_orders else 0,
                1 if has_offers else 0,
                1 if product.activo else 0,
                1 if has_primary_image else 0,
                1 if has_gallery else 0,
                1 if has_description else 0,
                1 if has_video else 0,
                1 if has_stock else 0,
                1 if has_price else 0,
                product.creado_en,
                product.pk,
            )

        return max(candidates, key=score)

    def _merge_same_name_duplicates(self, *, product, category_scoped=True):
        target = self._norm_compare_text(product.nombre)
        if not target:
            return product

        queryset = (
            Product.objects.exclude(pk=product.pk)
            .select_related("categoria")
            .prefetch_related("extra_images", "order_items", "ofertas")
        )
        if category_scoped:
            if product.categoria_id:
                queryset = queryset.filter(categoria_id=product.categoria_id)
            else:
                queryset = queryset.filter(categoria__isnull=True)

        duplicates = [
            candidate
            for candidate in queryset
            if self._norm_compare_text(candidate.nombre) == target
        ]
        if not duplicates:
            return product

        product = self._merge_products(
            survivor=product,
            duplicates=duplicates,
        )
        return product

    def _merge_products(self, *, survivor, duplicates):
        attrs = survivor.atributos if isinstance(survivor.atributos, dict) else {}
        attrs_stock = survivor.atributos_stock if isinstance(survivor.atributos_stock, dict) else {}
        attrs_price = survivor.atributos_precio if isinstance(survivor.atributos_precio, dict) else {}
        def gallery_key(image):
            if image.image_url:
                return ("url", image.image_url.strip())
            if image.image:
                return ("file", str(image.image))
            return ("id", image.pk)

        for current in sorted(duplicates, key=lambda item: (item.creado_en, item.pk)):
            ProductImage.objects.filter(product=current).update(product=survivor)
            Offer.objects.filter(producto=current).update(producto=survivor)
            current.order_items.update(product=survivor)

            is_newer = (current.creado_en, current.pk) > (survivor.creado_en, survivor.pk)

            if current.nombre and (is_newer or not survivor.nombre):
                survivor.nombre = current.nombre
            if current.descripcion and (is_newer or not survivor.descripcion):
                survivor.descripcion = current.descripcion
            if current.categoria_id and (is_newer or not survivor.categoria_id):
                survivor.categoria = current.categoria
            if current.image_url and (is_newer or not survivor.image_url):
                survivor.image_url = current.image_url
            elif getattr(current, "imagen", None) and not getattr(survivor, "imagen", None):
                survivor.imagen = current.imagen
            if current.video_url and (is_newer or not survivor.video_url):
                survivor.video_url = current.video_url
            if current.precio not in (None, "") and is_newer:
                survivor.precio = current.precio
            if current.stock is not None and is_newer:
                survivor.stock = current.stock
            survivor.activo = survivor.activo or current.activo

            current_attrs = current.atributos if isinstance(current.atributos, dict) else {}
            for key, values in current_attrs.items():
                existing = attrs.get(key, [])
                if not isinstance(existing, list):
                    existing = [existing] if existing else []
                incoming_values = values if isinstance(values, list) else [values]
                for value in incoming_values:
                    if value not in existing:
                        existing.append(value)
                attrs[key] = existing

            current_stock = current.atributos_stock if isinstance(current.atributos_stock, dict) else {}
            for key, value_map in current_stock.items():
                existing_map = attrs_stock.get(key, {}) if isinstance(attrs_stock.get(key), dict) else {}
                if isinstance(value_map, dict):
                    for value, stock_value in value_map.items():
                        existing_map[value] = max(int(existing_map.get(value, 0)), int(stock_value or 0))
                attrs_stock[key] = existing_map

            current_prices = current.atributos_precio if isinstance(current.atributos_precio, dict) else {}
            for key, value_map in current_prices.items():
                existing_map = attrs_price.get(key, {}) if isinstance(attrs_price.get(key), dict) else {}
                if isinstance(value_map, dict):
                    for value, price_value in value_map.items():
                        existing_map[value] = price_value
                attrs_price[key] = existing_map

        survivor.atributos = attrs if attrs else {}
        survivor.atributos_stock = attrs_stock if attrs_stock else {}
        survivor.atributos_precio = attrs_price if attrs_price else {}
        survivor.precio = self._resolve_base_price(survivor, fallback=survivor.precio)
        survivor.save()

        ordered_images = list(ProductImage.objects.filter(product=survivor).order_by("order", "id"))
        seen = set()
        next_order = 1
        for image in ordered_images:
            key = gallery_key(image)
            if key in seen:
                image.delete()
                continue
            seen.add(key)
            if image.order != next_order:
                image.order = next_order
                image.save(update_fields=["order"])
            next_order += 1

        duplicate_ids = [candidate.pk for candidate in duplicates]
        if duplicate_ids:
            Product.objects.filter(pk__in=duplicate_ids).delete()
        return survivor

    def _resolve_base_price(self, product, fallback):
        attrs = product.atributos if isinstance(product.atributos, dict) else {}
        price_map = product.atributos_precio if isinstance(product.atributos_precio, dict) else {}
        for attr_name, values in attrs.items():
            if not isinstance(values, list) or not values:
                continue
            by_attr = price_map.get(attr_name)
            if not isinstance(by_attr, dict):
                continue
            first_value = values[0]
            candidate = by_attr.get(first_value)
            if candidate not in (None, ""):
                return self._parse_decimal(candidate) or fallback
        return fallback

    def _build_group_slug(self, *, nombre, path_parts):
        base_slug = slugify(nombre or "")[:70] or "producto"
        raw = "|".join([
            self._norm_header(nombre),
            "|".join(self._norm_header(part) for part in (path_parts or [])),
        ])
        digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
        return f"{base_slug}-{digest}"[:110]

    def _build_export_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(EXPORT_HEADERS)
        for column_letter, width in EXPORT_COLUMN_WIDTHS.items():
            ws.column_dimensions[column_letter].width = width
        return wb

    def _load_export_base_workbook(self):
        if self.template_xlsx_path and os.path.isfile(self.template_xlsx_path):
            workbook = openpyxl.load_workbook(self.template_xlsx_path)
            ws = workbook.active
            headers = [ws.cell(row=1, column=index).value for index in range(1, len(EXPORT_HEADERS) + 1)]
            normalized_headers = [self._norm_header(header) for header in headers]
            normalized_export_headers = [self._norm_header(header) for header in EXPORT_HEADERS]
            if normalized_headers == normalized_export_headers:
                return workbook
        return None

    def _serialize_product_for_export(self, product):
        row = {header: "" for header in EXPORT_HEADERS}
        row["Nombre"] = product.nombre or ""
        row["Stock"] = product.stock if product.stock is not None else ""
        row["SKU"] = ""
        row["Precio"] = self._format_export_number(product.precio)
        row["Precio oferta"] = ""
        row["Categorías"] = self._build_export_category_path(product.categoria)
        row["Peso"] = ""
        row["Alto"] = ""
        row["Ancho"] = ""
        row["Profundidad"] = ""
        row["Mostrar en tienda"] = "Si" if product.activo else "No"
        row["IDProduct"] = product.pk or ""
        row["IDStock"] = ""

        image_urls = []
        primary_image = self._get_export_image_value(product.image_url, getattr(product, "imagen", None))
        if primary_image:
            image_urls.append(primary_image)
        for image in product.extra_images.all().order_by("order", "id"):
            image_value = self._get_export_image_value(image.image_url, getattr(image, "image", None))
            if image_value and image_value not in image_urls:
                image_urls.append(image_value)
        row["URL IMAGENES"] = " | ".join(image_urls)

        attrs = product.atributos if isinstance(product.atributos, dict) else {}
        for index, (name, values) in enumerate(attrs.items(), start=1):
            if index > 3:
                break
            row[f"Nombre atributo {index}"] = name
            if isinstance(values, list):
                row[f"Valor atributo {index}"] = ", ".join(str(value) for value in values if value not in (None, ""))
            elif values not in (None, ""):
                row[f"Valor atributo {index}"] = str(values)

        return row

    def _get_export_image_value(self, url_value, file_field):
        if url_value:
            return str(url_value).strip()
        if file_field:
            try:
                return file_field.url
            except Exception:
                return str(file_field)
        return ""

    def _build_export_category_path(self, category):
        if not category:
            return ""
        parts = []
        current = category
        while current:
            parts.append(current.nombre)
            current = current.parent
        return " > ".join(reversed(parts))

    def _format_export_number(self, value):
        if value in (None, ""):
            return ""
        decimal_value = self._parse_decimal(value)
        if decimal_value is None:
            return str(value)
        normalized = decimal_value.quantize(Decimal("0.01"))
        if normalized == normalized.to_integral():
            return str(int(normalized))
        return format(normalized.normalize(), "f")

    def _export_signature_parts(self, *, nombre="", categoria="", sku="", image_url=""):
        normalized_name = self._norm_header(nombre)
        normalized_category = self._norm_header(categoria)
        normalized_sku = self._norm_header(sku)
        normalized_url = self._norm_header(image_url)
        signatures = set()
        if normalized_name:
            signatures.add(("name", normalized_name))
        if normalized_name and normalized_category:
            signatures.add(("name_category", normalized_name, normalized_category))
        if normalized_name and normalized_category and normalized_url:
            signatures.add(("name_category_url", normalized_name, normalized_category, normalized_url))
        if normalized_sku:
            signatures.add(("sku", normalized_sku))
        if normalized_name and normalized_sku:
            signatures.add(("name_sku", normalized_name, normalized_sku))
        return signatures

    def _export_row_signatures(self, row):
        return self._export_signature_parts(
            nombre=row.get("Nombre", ""),
            categoria=row.get("Categorías", ""),
            sku=row.get("SKU", ""),
            image_url=row.get("URL IMAGENES", ""),
        )

    def _row_export_signatures(self, raw):
        row = {
            header: raw[index] if index < len(raw) else ""
            for index, header in enumerate(EXPORT_HEADERS)
        }
        return self._export_row_signatures(row)

