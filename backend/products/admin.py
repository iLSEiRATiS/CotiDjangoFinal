from decimal import Decimal
import hashlib
import os
import re
import openpyxl
from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.text import slugify
import unicodedata

from .models import Product, Category, Offer, HomeImage

admin.site.site_header = "Admin Coti"
admin.site.site_title = "Admin Coti"
admin.site.index_title = "Panel de administracion"


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("nombre", "slug", "parent")
    prepopulated_fields = {"slug": ("nombre",)}
    search_fields = ("nombre", "descripcion")
    list_filter = ("parent",)


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ("nombre", "slug", "precio", "user", "categoria", "stock", "activo", "creado_en")
    search_fields = ("nombre", "descripcion", "slug")
    list_filter = ("creado_en", "activo", "categoria")
    list_editable = ("precio", "stock", "activo")
    change_list_template = "admin/products/product/change_list.html"

    template_xlsx_path = r"C:\Users\facun\OneDrive\Escritorio\CotiWeb\Exportacion-productos-03-02-26.fixed.xlsx"

    product_headers = [
        "sku", "parent_sku", "nombre", "slug", "descripcion", "categoria", "subcategoria", "marca",
        "precio", "costo", "moneda", "stock", "activo", "opcion_1_nombre", "opcion_1_valor",
        "opcion_2_nombre", "opcion_2_valor", "imagen_1", "url imag", "imagen_2", "meta_title",
        "meta_description", "peso", "largo", "ancho", "alto", "es_destacado", "requiere_envio",
        "gestion_stock",
    ]

    sample_rows = [
        {
            "sku": "REM-BAS-001",
            "parent_sku": "",
            "nombre": "Remera basica",
            "slug": "remera-basica",
            "descripcion": "Remera algodon lisa",
            "categoria": "Ropa",
            "subcategoria": "Remeras",
            "marca": "Acme",
            "precio": 8999,
            "costo": 4500,
            "moneda": "ARS",
            "stock": 120,
            "activo": True,
            "opcion_1_nombre": "",
            "opcion_1_valor": "",
            "opcion_2_nombre": "",
            "opcion_2_valor": "",
            "imagen_1": "https://example.com/rem-bas-001.jpg",
            "url imag": "",
            "imagen_2": "",
            "meta_title": "Remera basica",
            "meta_description": "Remera de algodon basica",
            "peso": 0.3,
            "largo": 30,
            "ancho": 25,
            "alto": 2,
            "es_destacado": True,
            "requiere_envio": True,
            "gestion_stock": True,
        },
        {
            "sku": "REM-BAS-001-M-NEGRO",
            "parent_sku": "REM-BAS-001",
            "nombre": "Remera basica M Negro",
            "slug": "remera-basica-m-negro",
            "descripcion": "Remera algodon talla M color negro",
            "categoria": "Ropa",
            "subcategoria": "Remeras",
            "marca": "Acme",
            "precio": 8999,
            "costo": 4500,
            "moneda": "ARS",
            "stock": 40,
            "activo": True,
            "opcion_1_nombre": "Talle",
            "opcion_1_valor": "M",
            "opcion_2_nombre": "Color",
            "opcion_2_valor": "Negro",
            "imagen_1": "https://example.com/rem-bas-001-m-negro.jpg",
            "url imag": "",
            "imagen_2": "",
            "meta_title": "Remera basica M negro",
            "meta_description": "Remera negra basica talla M",
            "peso": 0.3,
            "largo": 30,
            "ancho": 25,
            "alto": 2,
            "es_destacado": False,
            "requiere_envio": True,
            "gestion_stock": True,
        },
    ]

    _KNOWN_COLORS = {
        "rojo": "Rojo",
        "azul": "Azul",
        "celeste": "Celeste",
        "verde": "Verde",
        "verde agua": "Verde Agua",
        "verde manzana": "Verde Manzana",
        "amarillo": "Amarillo",
        "naranja": "Naranja",
        "violeta": "Violeta",
        "morado": "Morado",
        "rosa": "Rosa",
        "rosa gold": "Rosa Gold",
        "fucsia": "Fucsia",
        "dorado": "Dorado",
        "oro": "Oro",
        "plateado": "Plateado",
        "plata": "Plata",
        "negro": "Negro",
        "blanco": "Blanco",
        "multicolor": "Multicolor",
        "turquesa": "Turquesa",
        "salmon": "Salmón",
        "calido": "Cálido",
    }

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "importar-xlsx/",
                self.admin_site.admin_view(self.import_xlsx_view),
                name="products_product_import_xlsx",
            ),
        ]
        return custom + urls

    def _parse_bool(self, value, default=True):
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        return s in {"true", "1", "si", "sí", "yes", "y"}

    def _parse_decimal(self, value):
        if value is None or value == "":
            return None
        s = str(value).replace("$", "").replace(" ", "").replace(",", ".")
        try:
            return Decimal(s)
        except Exception:
            return None

    def _parse_int(self, value):
        if value in (None, ""):
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    def _parse_attr_values(self, value):
        if value in (None, ""):
            return []
        text = str(value).strip()
        if not text:
            return []
        for sep in [",", ";", "|", "/"]:
            if sep in text:
                parts = [p.strip() for p in text.split(sep)]
                return [p for p in parts if p]
        return [text]

    def _strip_attr_from_name(self, name, values):
        if not name:
            return name
        base = str(name)
        for val in values or []:
            if not val:
                continue
            # remove value occurrences (case-insensitive)
            base = base.replace(str(val), "").replace(str(val).lower(), "").replace(str(val).upper(), "")
        # cleanup common separators
        base = base.replace("  ", " ").replace(" - ", " ").replace(" ,", ",").strip(" -_,")
        base = " ".join(base.split())
        return base or name

    def _build_slug(self, base):
        candidate = slugify(base or "")[:110] or "producto"
        original = candidate
        i = 1
        while Product.objects.filter(slug=candidate).exists():
            i += 1
            candidate = f"{original}-{i}"
        return candidate

    def _norm_header(self, value):
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(c for c in text if not unicodedata.combining(c))
        return text

    def _parse_category_path(self, raw):
        if not raw:
            return []
        text = str(raw).strip()
        if not text:
            return []
        # Soporta separadores jerarquicos sin romper nombres con guion interno
        # como "Miniaturas-Juguetitos" (sin espacios).
        for sep in (" > ", ">", " / ", "/", " | ", "|", " - ", " – "):
            if sep in text:
                parts = [p.strip() for p in text.split(sep)]
                return [p for p in parts if p]
        return [text]

    def _compose_category_path(self, categoria_raw, subcategoria_raw):
        categoria_parts = self._parse_category_path(categoria_raw)
        subcategoria_parts = self._parse_category_path(subcategoria_raw)
        if not categoria_parts and not subcategoria_parts:
            return []
        if not categoria_parts:
            return subcategoria_parts
        if not subcategoria_parts:
            return categoria_parts
        # Evita repetir nodos cuando subcategoria ya viene dentro de categoria.
        cat_norm = [self._norm_header(p) for p in categoria_parts]
        out = list(categoria_parts)
        for part in subcategoria_parts:
            if self._norm_header(part) not in cat_norm:
                out.append(part)
        return out

    def _get_or_create_category_normalized(self, name, parent=None):
        target = self._norm_header(name)
        for cat in Category.objects.filter(parent=parent):
            if self._norm_header(cat.nombre) == target:
                return cat
        return Category.objects.create(nombre=name, parent=parent)

    def _name_tokens(self, name):
        return re.findall(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9]+", str(name or ""))

    def _detect_color_in_name(self, name):
        n = self._norm_header(name).replace("/", " ")
        # prioriza match de frases largas
        for raw, val in sorted(self._KNOWN_COLORS.items(), key=lambda x: len(x[0]), reverse=True):
            if f" {raw} " in f" {n} ":
                return val
        return None

    def _detect_number_in_name(self, name):
        text = str(name or "")
        m = re.search(r"\bN\s*([0-9])\b", text, flags=re.IGNORECASE)
        if m:
            return m.group(1)
        m = re.search(r"\b([0-9])\b", text)
        if m:
            return m.group(1)
        return None

    def _common_root_name(self, a, b):
        wa = self._name_tokens(a)
        wb = self._name_tokens(b)
        i = 0
        while i < len(wa) and i < len(wb) and self._norm_header(wa[i]) == self._norm_header(wb[i]):
            i += 1
        if i < 2:
            return None
        root = " ".join(wa[:i]).strip()
        if not root:
            return None
        # Si alguno tiene "(ELEGIR ...)" se lo conservamos al nombre base.
        for src in (a, b):
            m = re.search(r"\((ELEGIR[^)]*)\)", str(src or ""), flags=re.IGNORECASE)
            if m and m.group(1):
                tag = m.group(1).strip()
                if tag and f"({tag})" not in root:
                    return f"{root} ({tag})"
        return root

    def _infer_attr_from_names(self, old_name, new_name):
        c_old = self._detect_color_in_name(old_name)
        c_new = self._detect_color_in_name(new_name)
        if c_new and c_new != c_old:
            return "Colores", c_new

        n_old = self._detect_number_in_name(old_name)
        n_new = self._detect_number_in_name(new_name)
        if n_new and n_new != n_old and ("elegir numero" in self._norm_header(old_name) or "elegir numero" in self._norm_header(new_name)):
            return "Número", n_new
        return None

    def _build_identity_slug(self, *, sku_raw, slug_raw, nombre, path_parts, precio, parent_sku_raw):
        sku_text = str(sku_raw or "").strip()
        if sku_text:
            return slugify(sku_text)[:110] or self._build_slug(nombre)

        base_slug = slugify(slug_raw or nombre or "")[:70] or "producto"
        path_key = "|".join(self._norm_header(p) for p in (path_parts or []))
        price_key = str(precio if precio is not None else "")
        raw = "|".join([
            self._norm_header(nombre),
            path_key,
            self._norm_header(parent_sku_raw),
            price_key,
        ])
        digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
        return f"{base_slug}-{digest}"[:110]

    def _build_group_slug(self, *, nombre, path_parts, precio):
        base_slug = slugify(nombre or "")[:70] or "producto"
        raw = "|".join([
            self._norm_header(nombre),
            "|".join(self._norm_header(p) for p in (path_parts or [])),
            str(precio if precio is not None else ""),
        ])
        digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
        return f"{base_slug}-{digest}"[:110]

    def _export_workbook(self, rows, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(self.product_headers)
        for row in rows:
            ws.append([row.get(h, "") for h in self.product_headers])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def import_xlsx_view(self, request):
        if request.method == "GET" and request.GET.get("sample"):
            return self._export_workbook(self.sample_rows, "productos_ejemplo.xlsx")
        if request.method == "GET" and request.GET.get("template"):
            if os.path.isfile(self.template_xlsx_path):
                with open(self.template_xlsx_path, "rb") as fh:
                    response = HttpResponse(
                        fh.read(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    response["Content-Disposition"] = 'attachment; filename="plantilla_productos.xlsx"'
                    return response
            empty_row = {h: "" for h in self.product_headers}
            return self._export_workbook([empty_row], "plantilla_productos.xlsx")

        created = 0
        updated = 0
        errors = []

        if request.method == "POST":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Selecciona un archivo XLSX.")
                return redirect(
                    reverse("admin:products_product_import_xlsx")
                )
            try:
                def _open_sheet():
                    upload.seek(0)
                    wb_local = openpyxl.load_workbook(upload, data_only=True, read_only=True)
                    return wb_local, wb_local.active

                wb, sheet = _open_sheet()
                header_idx = None
                header_row = None
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if not row:
                        continue
                    normalized = [self._norm_header(h) for h in row]
                    if {"sku", "nombre", "precio"}.issubset(set(normalized)):
                        header_idx = i
                        header_row = row
                        break
                wb.close()

                if header_row is None:
                    header_row = self.product_headers
                    header_idx = -1

                headers = [self._norm_header(h) for h in header_row]
                alias = {
                    "categorias": "categoria",
                    "categoria": "categoria",
                    "subcategoria": "subcategoria",
                    "sub categoria": "subcategoria",
                    "subcategor?a": "subcategoria",
                    "url imagenes": "url imag",
                    "url imag": "url imag",
                    "url_imag": "url imag",
                    "mostrar en tienda": "activo",
                    "nombre atributo 1": "opcion_1_nombre",
                    "valor atributo 1": "opcion_1_valor",
                    "nombre atributo1": "opcion_1_nombre",
                    "valor atributo1": "opcion_1_valor",
                    "nombre atributo 2": "opcion_2_nombre",
                    "valor atributo 2": "opcion_2_valor",
                    "nombre atributo2": "opcion_2_nombre",
                    "valor atributo2": "opcion_2_valor",
                }
                header_map = {}
                for idx, h in enumerate(headers):
                    key = alias.get(h, h)
                    if key:
                        header_map[key] = idx

                missing = {"sku", "nombre", "precio"} - set(header_map.keys())
                if missing:
                    errors.append(f"Faltan columnas obligatorias: {', '.join(sorted(missing))}")

                sku_name_sets = {}
                wb, sheet = _open_sheet()
                for row_idx, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if header_idx is not None and row_idx - 1 == header_idx:
                        continue
                    row_data = {
                        h: raw[header_map[h]] if h in header_map and header_map[h] < len(raw) else ""
                        for h in header_map
                    }
                    nombre_row = str(row_data.get("nombre") or "").strip()
                    sku_row = str(row_data.get("sku") or "").strip().upper()
                    if not nombre_row or not sku_row:
                        continue
                    sku_name_sets.setdefault(sku_row, set()).add(self._norm_header(nombre_row))
                wb.close()
                multi_name_skus = {sku for sku, names in sku_name_sets.items() if len(names) > 1}

                wb, sheet = _open_sheet()
                for idx, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if header_idx is not None and idx - 1 == header_idx:
                        continue
                    data = {h: raw[header_map[h]] if h in header_map and header_map[h] < len(raw) else "" for h in header_map}
                    row_data = {k: data.get(k) for k in header_map}
                    if all(v in ("", None) for v in row_data.values()):
                        continue
                    nombre = row_data.get("nombre") or ""
                    precio = self._parse_decimal(row_data.get("precio"))
                    if not nombre:
                        errors.append(f"Fila {idx}: nombre es obligatorio.")
                        continue
                    if precio is None:
                        precio = Decimal("0")
                    opt1_name = (row_data.get("opcion_1_nombre") or "").strip()
                    opt1_val = row_data.get("opcion_1_valor")
                    opt2_name = (row_data.get("opcion_2_nombre") or "").strip()
                    opt2_val = row_data.get("opcion_2_valor")
                    opt1_vals = self._parse_attr_values(opt1_val)
                    opt2_vals = self._parse_attr_values(opt2_val)
                    has_declared_attrs = bool((opt1_name and opt1_vals) or (opt2_name and opt2_vals))
                    sku_raw = row_data.get("sku") or ""
                    parent_sku_raw = row_data.get("parent_sku") or ""
                    slug_raw = row_data.get("slug") or ""
                    categoria_raw = row_data.get("categoria") or ""
                    subcategoria_raw = row_data.get("subcategoria") or ""
                    categoria_obj = None
                    path_parts = self._compose_category_path(categoria_raw, subcategoria_raw)
                    if path_parts:
                        parent = None
                        for part in path_parts:
                            parent = self._get_or_create_category_normalized(part, parent=parent)
                        categoria_obj = parent
                    if has_declared_attrs:
                        slug = self._build_group_slug(nombre=nombre, path_parts=path_parts, precio=precio)
                    else:
                        sku_upper = str(sku_raw).strip().upper()
                        effective_sku = "" if (sku_upper and sku_upper in multi_name_skus) else sku_raw
                        slug = self._build_identity_slug(
                            sku_raw=effective_sku,
                            slug_raw=slug_raw,
                            nombre=nombre,
                            path_parts=path_parts,
                            precio=precio,
                            parent_sku_raw=parent_sku_raw or sku_raw,
                        )
                    existing = Product.objects.filter(slug=slug).first()
                    is_new = existing is None
                    product = existing or Product(slug=slug, user=request.user)
                    product.nombre = nombre
                    product.descripcion = row_data.get("descripcion") or ""
                    product.precio = precio
                    stock = self._parse_int(row_data.get("stock"))
                    product.stock = stock if stock is not None else 0
                    product.activo = self._parse_bool(row_data.get("activo"), default=True)
                    product.categoria = categoria_obj
                    attrs = {}
                    if opt1_name and opt1_vals:
                        attrs[opt1_name] = opt1_vals
                    if opt2_name and opt2_vals:
                        attrs[opt2_name] = opt2_vals
                    has_sku = bool(str(sku_raw).strip())
                    if has_declared_attrs:
                        merged = {}
                        if isinstance(product.atributos, dict):
                            merged.update(product.atributos)
                        for key, values in attrs.items():
                            current = merged.get(key, [])
                            if not isinstance(current, list):
                                current = [current] if current else []
                            for v in values:
                                if v not in current:
                                    current.append(v)
                            merged[key] = current
                        product.atributos = merged if merged else {}
                        stock_map = product.atributos_stock if isinstance(product.atributos_stock, dict) else {}
                        if stock is not None:
                            for key, values in attrs.items():
                                existing_map = stock_map.get(key, {}) if isinstance(stock_map.get(key), dict) else {}
                                for v in values:
                                    existing_map[v] = max(int(existing_map.get(v, 0)), int(stock))
                                stock_map[key] = existing_map
                        product.atributos_stock = stock_map if stock_map else {}
                    elif has_sku:
                        product.atributos = attrs if attrs else {}
                        if stock is not None and attrs:
                            stock_map = {}
                            for key, values in attrs.items():
                                stock_map[key] = {v: int(stock) for v in values}
                            product.atributos_stock = stock_map
                        elif not attrs:
                            product.atributos_stock = {}
                    else:
                        product.atributos = {}
                        product.atributos_stock = {}
                    image_url = row_data.get("imagen_1") or row_data.get("url imag") or row_data.get("url_imag") or ""
                    if image_url and str(image_url).startswith(("http://", "https://")):
                        product.image_url = str(image_url).strip()
                    if is_new and not product.slug:
                        product.slug = self._build_slug(nombre)
                    product.save()
                    if is_new:
                        created += 1
                    else:
                        updated += 1
                wb.close()
                if created or updated:
                    messages.success(
                        request,
                        f"Importaci?n completada. Nuevos: {created} | Actualizados: {updated}",
                    )
                if errors:
                    for err in errors:
                        messages.error(request, err)
            except Exception as exc:  # pragma: no cover
                messages.error(request, f"No se pudo procesar el XLSX: {exc}")
                return redirect(reverse("admin:products_product_import_xlsx"))

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Importar productos via XLSX",
            "headers": self.product_headers,
            "example_url": f"{reverse('admin:products_product_import_xlsx')}?sample=1",
            "template_url": f"{reverse('admin:products_product_import_xlsx')}?template=1",
            "created": created,
            "updated": updated,
            "errors": errors,
        }
        return TemplateResponse(request, "admin/products/product/import_xlsx.html", context)



@admin.register(Offer)
class OfferAdmin(admin.ModelAdmin):
    list_display = ("nombre", "porcentaje", "producto", "categoria", "activo", "empieza", "termina")
    list_filter = ("activo", "empieza", "termina", "categoria")
    search_fields = ("nombre", "descripcion", "producto__nombre", "categoria__nombre")
    prepopulated_fields = {"slug": ("nombre",)}
    list_editable = ("activo",)
    actions = ["activar_ofertas", "desactivar_ofertas"]

    @admin.action(description="Activar ofertas seleccionadas")
    def activar_ofertas(self, request, queryset):
        queryset.update(activo=True)

    @admin.action(description="Desactivar ofertas seleccionadas")
    def desactivar_ofertas(self, request, queryset):
        queryset.update(activo=False)


@admin.register(HomeImage)
class HomeImageAdmin(admin.ModelAdmin):
    list_display = ("key", "section", "title", "target_url", "order", "activo")
    list_filter = ("section", "activo")
    search_fields = ("key", "title", "image_url", "target_url")
    list_editable = ("order", "activo")
