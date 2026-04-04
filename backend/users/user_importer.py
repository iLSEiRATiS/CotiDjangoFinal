import os

import openpyxl
from django.http import HttpResponse

from .models import CustomUser


USER_IMPORT_HEADERS = [
    "nombre",
    "apellido",
    "email",
    "password",
    "approval_status",
    "phone",
    "address",
    "city",
    "zip_code",
]


USER_SAMPLE_ROWS = [
    {
        "nombre": "Valentin",
        "apellido": "Geres",
        "email": "valentin@example.com",
        "password": "Cliente123!",
        "approval_status": "approved",
        "phone": "",
        "address": "",
        "city": "",
        "zip_code": "",
    }
]


HEADER_ALIAS = {
    "nombre": "nombre",
    "name": "nombre",
    "first_name": "nombre",
    "apellido": "apellido",
    "last_name": "apellido",
    "email": "email",
    "correo": "email",
    "password": "password",
    "clave": "password",
    "contrasena": "password",
    "approval_status": "approval_status",
    "estado de aprobacion": "approval_status",
    "phone": "phone",
    "telefono": "phone",
    "address": "address",
    "direccion": "address",
    "city": "city",
    "ciudad": "city",
    "zip_code": "zip_code",
    "cp": "zip_code",
    "codigo postal": "zip_code",
}


class UserXlsxImporter:
    def __init__(self, *, template_xlsx_path=None):
        self.template_xlsx_path = template_xlsx_path

    def export_workbook(self, rows, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        ws.append(USER_IMPORT_HEADERS)
        for row in rows:
            ws.append([row.get(header, "") for header in USER_IMPORT_HEADERS])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_template_response(self):
        if self.template_xlsx_path and os.path.isfile(self.template_xlsx_path):
            with open(self.template_xlsx_path, "rb") as fh:
                response = HttpResponse(
                    fh.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = 'attachment; filename="plantilla_clientes.xlsx"'
                return response
        return self.export_workbook([{header: "" for header in USER_IMPORT_HEADERS}], "plantilla_clientes.xlsx")

    def import_upload(self, upload):
        wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return 0, 0, ["El archivo no contiene filas."]

        headers = [self._norm_header(value) for value in rows[0]]
        mapped = {}
        for idx, header in enumerate(headers):
            canonical = HEADER_ALIAS.get(header)
            if canonical:
                mapped[canonical] = idx

        required = {"nombre", "apellido", "email"}
        missing = required - set(mapped.keys())
        if missing:
            return 0, 0, [f"Faltan columnas obligatorias: {', '.join(sorted(missing))}"]

        created = 0
        updated = 0
        errors = []

        for row_number, raw in enumerate(rows[1:], start=2):
            data = {
                key: raw[idx] if idx < len(raw) else ""
                for key, idx in mapped.items()
            }
            if all(value in ("", None) for value in data.values()):
                continue

            first_name = str(data.get("nombre") or "").strip()
            last_name = str(data.get("apellido") or "").strip()
            email = str(data.get("email") or "").strip().lower()
            password = str(data.get("password") or "").strip()
            approval_status = str(data.get("approval_status") or "pending").strip().lower() or "pending"
            phone = str(data.get("phone") or "").strip()
            address = str(data.get("address") or "").strip()
            city = str(data.get("city") or "").strip()
            zip_code = str(data.get("zip_code") or "").strip()

            if not first_name or not last_name or not email:
                errors.append(f"Fila {row_number}: nombre, apellido y email son obligatorios.")
                continue
            if approval_status not in {"pending", "approved", "rejected"}:
                errors.append(f"Fila {row_number}: estado de aprobacion invalido.")
                continue

            user = CustomUser.objects.filter(email__iexact=email).first()
            is_new = user is None
            if is_new:
                user = CustomUser(username=email, email=email)
            user.username = email
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.name = f"{first_name} {last_name}".strip()
            user.phone = phone
            user.address = address
            user.city = city
            user.zip_code = zip_code
            user.approval_status = approval_status
            if password:
                user.set_password(password)
            elif is_new:
                user.set_unusable_password()
            user.save()

            if is_new:
                created += 1
            else:
                updated += 1

        return created, updated, errors

    def _norm_header(self, value):
        return str(value or "").strip().lower()
